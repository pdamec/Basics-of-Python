"""Module4/Task4: Spreadsheet Cell Inverter."""
import os

import openpyxl


# Vars
SPREADSHEET_PATH = 'resources/task_4/dummy_1.xlsx'


def invert_spreadsheet_cells(spreadsheet: str):
    """Invert cells in a spreadsheet and save it in current directory.

    :param spreadsheet: path to the spreadsheet.
    """
    current_wb = openpyxl.load_workbook(spreadsheet)
    current_sheet = current_wb.active

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    for row in range(1, current_sheet.max_row + 1):
        for col in range(1, current_sheet.max_column + 1):
            new_sheet.cell(col, row).value = current_sheet.cell(row, col).value

    filename = f'inverted_{os.path.basename(spreadsheet)}'
    new_wb.save(filename)


if __name__ == '__main__':
    invert_spreadsheet_cells(SPREADSHEET_PATH)
