"""Module4/Task7: Excel to CSV converter."""
import os
import csv

from openpyxl import load_workbook

from tasks.module_4.task_2 import cwd


# Vars
EXCEL_DIR = 'resources/task_7'
EXCEL_EXTENSIONS = ['.xlsx']


def get_excel_files() -> list:
    """Get list of excel files in current directory."""
    files = []
    for file_ in os.listdir():
        if os.path.isfile(file_) and os.path.splitext(file_)[1] in EXCEL_EXTENSIONS:
            files.append(file_)
    return files


def convert_excel_to_csv(file: str):
    """Create csv files based on excel sheet.

    One sheet equals one file.

    Naming:
    <excel filename>_<sheet title>.csv, where <excel filename>
    is the filename of the Excel file without the file extension
    (for example, 'spam_data', not 'spam_data.xlsx') and <sheet title>
    is the string from the Worksheet object’s title variable.

    :param file: path to the excel sheet.
    """
    work_book = load_workbook(file)

    for sheet in work_book.sheetnames:
        sheet = work_book[sheet]

        filename = f'{os.path.splitext(file)[0]}_{sheet.title}.csv'

        with open(filename, 'w', newline="") as f_out:
            csv_out = csv.writer(f_out)
            for row in sheet.rows:
                csv_out.writerow([cell.value for cell in row])


if __name__ == "__main__":
    with cwd(EXCEL_DIR):
        for excel_file in get_excel_files():  # Loop in function's scope?
            convert_excel_to_csv(excel_file)
